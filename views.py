from django.shortcuts import render
from django.http import HttpResponse
from .models import machine, svr, windowsSvrLicense, windowsLicense, officeLicense, term, termShift
from .models import machine_form,svr_form,windowsSvrLicense_form,officeLicense_form, termShift_form,term_form
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404
from apps.inventory.functions import *
from django.db.models import Sum
from django.views.decorators.csrf import csrf_exempt
import os, sys, openpyxl, datetime,time
from django.contrib.auth.decorators import login_required
# Create your views here.
def summary(request):
    SumFree = officeLicense.objects.all().aggregate(Sum('free'))['free__sum']
    SumAvailable = officeLicense.objects.all().aggregate(Sum('available'))['available__sum']
    SumInUse = officeLicense.objects.all().aggregate(Sum('inUse'))['inUse__sum']
    return render(request,"inventory/summary.html",locals())
def winsvrlic(request):
    licenses = windowsSvrLicense.objects.all()
    return render(request,"inventory/winSrvLic.html",locals())
def officelic(request):
    licenses = officeLicense.objects.all()
    return render(request,"inventory/officeLic.html",locals())

#@login_required(login_url='inventory/accounts/login/')
def workstations(request):
    filtr ="(|(cn=jan-pc-*)(cn=jan-nb-*)(cn=jan-tab-*)(cn=jan-szw*)(cn=jan-atm*))"
    #filtr ="(|(cn=jan-pc-*)(cn=jan-nb-*)(cn=jan-tab-*)(cn=jan-szw*)(cn=jan-atm*)(ou=IT))"
    #filtr ="(&(cn=jan-pc-*)(!(ou:dn:=Dark*)))"
    #filtr ="(cn=jan-pc-085)"
    #filtr ="(&(cn=jan-pc-*)(cn=jan-nb-*)(ou='Dark side'))"
    check_credentials('inventory','G7XZ:]+7J3*$wz!.',filtr,'pc')
    workstations = machine.objects.exclude(OU="Dark side").exclude(OU="Niekatywne")
    return render(request,"inventory/workstations_html.html",locals())
def servers(request):
    #filtr ="(cn=jan-svr-*)"
    #filtr = "(ou=Dark*)"
    #filtr ="(&(cn=jan-svr-*)(|(ou:dn:=SERVER)(ou:dn:=Computers))(!(cn=*-pr-*)))"
    filtr ="(&(cn=jan-svr-*)(!(cn=*-pr-*)))"
    #filtr ="(&(cn=jan-svr-*)(|(ou:dn:=SERVER)(ou:dn:=Computers)))"

    check_credentials('inventory','G7XZ:]+7J3*$wz!.',filtr,'svr')
    svrs = svr.objects.exclude(OU="Dark side").exclude(OU="Niekatywne")
    return render(request,"inventory/svrs.html",locals())
def terminals(request):
    terms = term.objects.all()
    return render(request,"inventory/terms.html",locals())
def terms_del(request):
    url = 'inventory/terms_del.html'
    #term_shifts.objects.filter()
    #TermsDel = termShift.objects.filter(description__contains="yliza")
    TermsDel.append =  termShift.objects.filter(typ=5)
    #TermsDel = termShift.objects.exclude(description__contains="szkodz",typ=5)

    context = {
    }
    context['TermsDel'] = TermsDel
    
    
    return render(request,url,context)
def winsvrlic_new(request):
    if request.method=='POST':
        form = windowsSvrLicense_form(request.POST)
        if form.is_valid():
            license = windowsSvrLicense()
            license.key = form.cleaned_data["key"]
            license.available = form.cleaned_data["available"]
            license.osName = form.cleaned_data["osName"]
            license.description = form.cleaned_data["description"]
            license.free = form.cleaned_data["free"]
            license.inUse = form.cleaned_data["inUse"]
        
            #mach.cn = request.cn
            license.save()
        else:
            print("BLAD")
    else:
        form = windowsSvrLicense_form()
        print("OK")
    return render(request,"inventory/winSvrLic_new.html",locals())

def officelic_new(request):
    if request.method=='POST':
        form = officeLicense_form(request.POST)
        if form.is_valid():
            license = officeLicense()
            license.key = form.cleaned_data["key"]
            license.available = form.cleaned_data["available"]
            license.description = form.cleaned_data["description"]
            license.free = form.cleaned_data["free"]
            license.inUse = form.cleaned_data["inUse"]
        
            #mach.cn = request.cn
            license.save()
        else:
            print("BLAD")
    else:
        form = officeLicense_form()
        print("OK")
    return render(request,"inventory/officeLic_new.html",locals())

def winsvrlic_del(request,pk):
    lic = windowsSvrLicense.objects.filter(pk=pk).delete()
    return render(request,"inventory/winSrvLic.html",locals())
    
def winsvrlic_edit(request,pk):
    license = windowsSvrLicense.objects.get(pk=pk)
    if request.method=='POST':
        form = windowsSvrLicense_form(request.POST, instance = license)
        if form.is_valid(): 
            #mach.cn = request.cn
            license.save()
        
        else:
            print("BLAD")
    else:
        None
       # form = windowsSvrLicense_form(instance = license)
    return render(request,"inventory/winSvrLic_edit.html",locals())
def workstation_edit(request,pk):
    nic = officeLicense
    mach = machine.objects.get(pk=pk)
    officeLicOld = mach.officeLicense
    users = user.objects.all()
    officeLicenses = officeLicense.objects.filter(free=1)
    print(officeLicenses)
    if request.method=='POST':
        form = machine_form(request.POST, instance = mach)
        if form.is_valid():
            mach.save()
            try:
                officeLicNew = officeLicense.objects.filter(key=mach.officeLicense.key)[0]
                if officeLicNew!=officeLicOld:
                    officeLicNew.free -=1 
                    officeLicNew.inUse +=1 
                    officeLicNew.save()
                    if officeLicOld!=None:
                        officeLicOld.free +=1
                        officeLicOld.inUse -=1
                        officeLicOld.save()
            except:
                
                    if officeLicOld!=None:
                        officeLicOld.free +=1
                        officeLicOld.inUse -=1
                        officeLicOld.save()
        else:
            print("BLAD----")


    else:
        officeLicOld = mach.officeLicense
        form = machine_form(instance = mach)
        print("------",mach.aktualizacja)
        print("OK")
    return render(request,"inventory/workstation_edit.html",locals())
def svr_edit(request,pk):
    nic = windowsSvrLicense
    svr1 = svr.objects.get(pk=pk)
    WinLicOld = svr1.windowsSvrLicense
    users = user.objects.all()
    licenses = windowsSvrLicense.objects.all()
    if request.method=='POST':
        form = svr_form(request.POST, instance = svr1)
        if form.is_valid():
            svr1.save()
            try:
                WinLicNew = officeLicense.objects.filter(key=svr1.WindowsLicense.key)[0]
                if WinLicNew!=WinLicOld:
                    WinLicNew.free -=1 
                    WinLicNew.inUse +=1 
                    WinLicNew.save()
                    if WinLicOld!=None:
                        WinLicOld.free +=1
                        WinLicOld.inUse -=1
                        WinLicOld.save()
            except:
                    WinLicOld.free +=1
                    WinLicOld.inUse -=1
                    WinLicOld.save()
        else:
            print("BLAD1111")
    else:
        WinLicOld = svr1.windowsSvrLicense
        form = svr_form(instance = svr1)
        print("OK")
    return render(request,"inventory/svr_edit.html",locals())
def users(request):
    wynik = check_credentials_users('inventory','G7XZ:]+7J3*$wz!.')
    return HttpResponse(wynik)

def officelic_edit(request,pk):
    license = officeLicense.objects.get(pk=pk)
    if request.method =='POST':
        form = officeLicense_form(request.POST, instance = license)
        if form.is_valid():
            license.save()
        else:
            None
 
    return render(request,"inventory/officeLic_edit.html",locals())
def term_edit(request,pk):
    trm = term.objects.get(pk=pk)
    print(trm)
    if request.method =='POST':
        form = term_form(request.POST, instance = trm)
        if form.is_valid():
            trm.save()
        else:
            None
    return render(request,"inventory/term_edit.html",locals())

def vncview(request):
    #os.popen('java -jar tika-app.jar -t %s > textOutput.txt' % fileName)    
    os.popen('java -jar tightvnc-jviewer.jar')

    return render(request,"inventory/vncview.html",locals())
    #return render(request,"inventory/summary.html",locals())
def spot(request):
    #Message table 
    #,[MessageType],[TimeStamp],[TimeClient],[TimeManual],[MachineId],[TerminalId],[WorkCenterId],[EmployeeId]
    #,[TimeCardNr],[OrderNr],[OperationNr],[ConfirmationNr],[Plant],[Shift],[WageGroup],[IsProcessed] 
    readMessages()
    readFoam()
    FulfillFoam()


    return render(request,"inventory/spot.html",locals())
@csrf_exempt
def kj(request):
    etykieta = request.POST['etykieta']
    print(etykieta)
    return render(request,"inventory/spot.html",locals())
#TERMINALE
def term_import(request):
    wb = openpyxl.load_workbook("apps/inventory/tmp/skanery.xlsx")
    ws = wb.active
    range2 = 'I'+str(ws.max_row)
    arkusz = ws['A2':range2]
    for hostname, sn, model, description, dep, soft, doWymiany, historia, mac in arkusz:
        try:
            if sn.value != '.':
                t = term.objects.get(sn=sn.value)
                t.hostname = hostname.value 
                t.sn = sn.value
                t.model = model.value
                t.description = description.value
                t.dep = dep.value
                t.mac = mac.value
                t.save()
        except:
            t = term.objects.create(hostname=hostname.value,sn=sn.value,model=model.value,description=description.value,dep=dep.value,mac=mac.value)
        print(t)
            
        

    return HttpResponse("OK")

def delete_unrelated_users(request):
    delete_unrelated()
    return HttpResponse("OK")
def workstations_export(request):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "WorkStations"
    workstations = machine.objects.all()
    row=1
    for workstation in workstations:
        cell = "A" + str(row)
        ws[cell] = workstation.cn
        cell = "B" + str(row)
        ws[cell] = workstation.operatingSystem
        cell = "C" + str(row)
        ws[cell] = workstation.description
        cell = "D" + str(row)
        ws[cell] = workstation.distinguishedName
        cell = "E" + str(row)
        ws[cell] = workstation.OU

        cell = "F" + str(row)
        ws[cell] =workstation.dNSHostName
        cell = "G" + str(row)
        try:
            ws[cell] = workstation.user.cn
        except:
            None
        cell = "H" + str(row)
        try:
            ws[cell] = workstation.windowsLicense.key
        except:
            None

        cell = "I" + str(row)
        try:
            ws[cell] = workstation.officeLicense.key
        except:
        
            None

        cell = "J" + str(row)
        ws[cell] =workstation.rodo


        row+=1

    wb.save('workstations-1.xlsx')
        

    return HttpResponse("OK")


def eventForm(request, *arg, **kwargs):
    url = 'inventory/form_terminal.html'
    context = {
    }
    if kwargs.get('event'):            
        event = kwargs.get('event')
        skaner = term.objects.get(pk=event)
        form = termShift_form()
    else:
        return HttpResponse("404 Error")
    if request.POST:
        form = termShift_form(request.POST)
        if form.is_valid():
            form.term = skaner            
            form.date = datetime.datetime.now()
            form.save()
        context['form'] = form        
        return render(request, url, context)       
    context['form'] = form
    return render(request, url, context)

def newEventForm(request, *arg, **kwargs):
    url = 'inventory/termShift_new.html'
    if kwargs.get('term'):
        pk = kwargs.get('term')
        terminal = term.objects.get(pk=pk)
        
    else:
        return("404 Error")

    if request.method=='POST':
        form = termShift_form(request.POST)
        if form.is_valid():
            termSh = termShift()
            termSh.term = terminal
            termSh.eventDate = form.cleaned_data["eventDate"]
            termSh.hostnameBefore  = form.cleaned_data["hostnameBefore"]
            termSh.hostnameAfter  = form.cleaned_data["hostnameAfter"]
            termSh.depBefore  = form.cleaned_data["depBefore"]
            termSh.depAfter  = form.cleaned_data["depAfter"]
            termSh.odbierajacy  = form.cleaned_data["odbierajacy"]
            termSh.wydajacy  = form.cleaned_data["wydajacy"]
            termSh.description  = form.cleaned_data["description"]
            termSh.typ  = form.cleaned_data["typ"]
            termSh.save()
        else:
            print("Inconsistent data error")
    else:
        form = termShift_form()
        form.eventDate = datetime.datetime.now()
    return render(request,url,locals())
def eventFilter(request, *arg, **kwargs):
    url = 'inventory/event_list.html'
    context = {
    }
    if kwargs.get('term'):
        termPk = kwargs.get('term')
        termShifts = termShift.objects.filter(term=termPk)
    
    else:
        termShifts = termShift.objects.all()
    return render(request,url,locals())

